"""Motore Fantacalcio Mantra - porting Python della logica VBA di Asta_Mantra.xlsm.

Backend puro (nessuna dipendenza da Streamlit). Carica i dati di riferimento dal
workbook (Listone / Moduli / Squadre) e replica l'ottimizzatore di formazione
`RisolviSlot` (assegnamento ottimo dei giocatori agli 11 slot di un modulo tramite
programmazione dinamica su bitmask), la classifica dei moduli, la formazione tipo
e la tabella di scarsita' per ruolo.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

import openpyxl

# ----------------------------------------------------------------------------
# Costanti (equivalenti alle Const del modulo VBA)
# ----------------------------------------------------------------------------
RUOLI: list[str] = ["Por", "Dd", "Ds", "Dc", "B", "E", "M", "C", "W", "T", "A", "Pc"]
ROLE_INDEX: dict[str, int] = {r: i for i, r in enumerate(RUOLI)}
NUM_RUOLI = len(RUOLI)  # 12

NSLOT = 11
CREDITI = 500
MIN_POR = 3
MAX_ROSA = 80
BONUS = 100.0  # premia la copertura di uno slot in piu' rispetto al rating

# Posizioni (1-based) nel foglio Listone
_LIS_ROW0, _LIS_ROW1 = 2, 540
_LIS_COL_NOME = 2
_LIS_COL_SQ = 3
_LIS_COL_FUORI = 7
_LIS_COL_RATING0 = 17  # Rt_Por .. Rt_Pc occupano 17..28

# Foglio Squadre
_SQ_ROW0, _SQ_ROW1 = 4, 11
_SQ_COL_NOME = 2
_SQ_COL_MIA = 3

# Foglio Moduli
_MOD_ROW0, _MOD_ROW1 = 5, 40
_MOD_COL_NOME = 1
_MOD_COL_SLOT0 = 2  # Slot 1 .. Slot 11 -> colonne 2..12


# ----------------------------------------------------------------------------
# Modelli dati
# ----------------------------------------------------------------------------
@dataclass
class Player:
    idx: int
    nome: str
    squadra_reale: str
    fuori_lista: bool
    ratings: list[float]  # lunghezza 12, -1.0 se il giocatore non copre quel ruolo

    def best_role_for_slot(self, slot_roles: list[int]) -> tuple[int, float]:
        """Ruolo migliore (indice, rating) fra quelli ammessi dallo slot; (-1, -1) se nessuno."""
        best_role, best_val = -1, -1.0
        for role in slot_roles:
            v = self.ratings[role]
            if v > best_val:
                best_val, best_role = v, role
        return best_role, best_val


@dataclass
class ReferenceData:
    players: list[Player]
    by_name: dict[str, int]          # nome (lower) -> idx nel listone
    modules: dict[str, list[list[int]]]  # nome modulo -> 11 slot, ognuno lista di indici ruolo
    module_order: list[str]
    teams: list[str]
    my_team: str

    def canonical_name(self, name: str) -> Optional[str]:
        idx = self.by_name.get(name.strip().lower())
        return None if idx is None else self.players[idx].nome


# ----------------------------------------------------------------------------
# Caricamento dati di riferimento
# ----------------------------------------------------------------------------
def _to_rating(value) -> float:
    if value is None:
        return -1.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return -1.0


def load_reference(path: str = "Asta_Mantra.xlsm") -> ReferenceData:
    """Legge Listone, Moduli e Squadre dal workbook e costruisce ReferenceData."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        players, by_name = _load_players(wb["Listone"])
        modules, module_order = _load_modules(wb["Moduli"])
        teams, my_team = _load_teams(wb["Squadre"])
    finally:
        wb.close()
    return ReferenceData(
        players=players,
        by_name=by_name,
        modules=modules,
        module_order=module_order,
        teams=teams,
        my_team=my_team,
    )


def _load_players(ws) -> tuple[list[Player], dict[str, int]]:
    players: list[Player] = []
    by_name: dict[str, int] = {}
    for r in range(_LIS_ROW0, _LIS_ROW1 + 1):
        nome = ws.cell(r, _LIS_COL_NOME).value
        if nome is None or str(nome).strip() == "":
            continue
        nome = str(nome).strip()
        squadra = str(ws.cell(r, _LIS_COL_SQ).value or "").strip()
        fuori = str(ws.cell(r, _LIS_COL_FUORI).value or "").strip() != ""
        ratings = [
            _to_rating(ws.cell(r, _LIS_COL_RATING0 + j).value) for j in range(NUM_RUOLI)
        ]
        idx = len(players)
        players.append(
            Player(idx=idx, nome=nome, squadra_reale=squadra, fuori_lista=fuori, ratings=ratings)
        )
        key = nome.lower()
        if key not in by_name:  # primo che vince, come dLis nel VBA
            by_name[key] = idx
    return players, by_name


def _load_modules(ws) -> tuple[dict[str, list[list[int]]], list[str]]:
    modules: dict[str, list[list[int]]] = {}
    order: list[str] = []
    for r in range(_MOD_ROW0, _MOD_ROW1 + 1):
        nome = ws.cell(r, _MOD_COL_NOME).value
        if nome is None or str(nome).strip() == "":
            break  # NumModuli si ferma alla prima riga vuota
        nome = str(nome).strip()
        slots: list[list[int]] = []
        for s in range(NSLOT):
            raw = str(ws.cell(r, _MOD_COL_SLOT0 + s).value or "").strip()
            roles = [ROLE_INDEX[p.strip()] for p in raw.split("/") if p.strip() in ROLE_INDEX]
            slots.append(roles)
        modules[nome] = slots
        order.append(nome)
    return modules, order


def _load_teams(ws) -> tuple[list[str], str]:
    teams: list[str] = []
    my_team = ""
    for r in range(_SQ_ROW0, _SQ_ROW1 + 1):
        nome = ws.cell(r, _SQ_COL_NOME).value
        if nome is None or str(nome).strip() == "":
            continue
        nome = str(nome).strip()
        teams.append(nome)
        if str(ws.cell(r, _SQ_COL_MIA).value or "").strip().upper() == "X" and not my_team:
            my_team = nome
    if not my_team and teams:
        my_team = teams[0]
    return teams, my_team


# ----------------------------------------------------------------------------
# Ottimizzatore di formazione (porting di RisolviSlot)
# ----------------------------------------------------------------------------
@dataclass
class SlotSolution:
    assignment: list[Optional[int]]  # per slot: idx giocatore o None
    roles: list[Optional[int]]       # per slot: indice ruolo usato o None
    points: list[float]              # per slot: rating usato
    total: float                     # somma dei rating (bonus escluso)

    @property
    def covered(self) -> int:
        return sum(1 for a in self.assignment if a is not None)

    @property
    def uncovered(self) -> int:
        return NSLOT - self.covered


def solve_slots(player_idxs: list[int], module_slots: list[list[int]], players: list[Player]) -> SlotSolution:
    """Assegnamento ottimo dei giocatori sugli 11 slot del modulo.

    Ogni accoppiamento vale rating + BONUS: coprire uno slot in piu' batte sempre un
    rating migliore su meno slot. Il bonus viene poi tolto dal totale restituito.
    Porting fedele del DP su bitmask del VBA.
    """
    empty = SlotSolution(
        assignment=[None] * NSLOT, roles=[None] * NSLOT, points=[0.0] * NSLOT, total=0.0
    )
    nP = len(player_idxs)
    if nP < 1:
        return empty

    # sc[i][s] = miglior rating del giocatore i sullo slot s (-1 se non lo copre)
    sc: list[list[float]] = [[-1.0] * NSLOT for _ in range(nP)]
    eligible: list[list[int]] = [[] for _ in range(nP)]
    for i, pidx in enumerate(player_idxs):
        rts = players[pidx].ratings
        for s in range(NSLOT):
            best = -1.0
            for role in module_slots[s]:
                v = rts[role]
                if v > best:
                    best = v
            sc[i][s] = best
            if best >= 0:
                eligible[i].append(s)

    n_mask = 1 << NSLOT  # 2048
    NEG = float("-inf")
    # dp[i][mask] = miglior valore usando i primi i giocatori con gli slot in mask coperti
    dp = [[NEG] * n_mask for _ in range(nP + 1)]
    choice = [[-2] * n_mask for _ in range(nP + 1)]  # -2 non raggiunto, -1 skip, >=0 slot scelto
    dp[0][0] = 0.0

    for i in range(nP):
        cur, nxt = dp[i], dp[i + 1]
        nch = choice[i + 1]
        el, sci = eligible[i], sc[i]
        for mask in range(n_mask):
            val = cur[mask]
            if val == NEG:
                continue
            # opzione 1: non schierare il giocatore i
            if val > nxt[mask]:
                nxt[mask] = val
                nch[mask] = -1
            # opzione 2: assegnarlo a uno slot libero ammesso
            for s in el:
                bit = 1 << s
                if not (mask & bit):
                    nm = mask | bit
                    nv = val + sci[s] + BONUS
                    if nv > nxt[nm]:
                        nxt[nm] = nv
                        nch[nm] = s

    # miglior stato finale
    fin = dp[nP]
    best_mask, best_val = 0, NEG
    for mask in range(n_mask):
        if fin[mask] > best_val:
            best_val, best_mask = fin[mask], mask

    # backtrack
    assignment: list[Optional[int]] = [None] * NSLOT
    roles: list[Optional[int]] = [None] * NSLOT
    points = [0.0] * NSLOT
    total = 0.0
    mask = best_mask
    for i in range(nP, 0, -1):
        s = choice[i][mask]
        if s is not None and s >= 0:
            pidx = player_idxs[i - 1]
            assignment[s] = pidx
            points[s] = sc[i - 1][s]
            total += sc[i - 1][s]
            role, _ = players[pidx].best_role_for_slot(module_slots[s])
            roles[s] = role
            mask &= ~(1 << s)

    return SlotSolution(assignment=assignment, roles=roles, points=points, total=total)


# ----------------------------------------------------------------------------
# Stato della lega derivato dagli acquisti
# ----------------------------------------------------------------------------
@dataclass
class LeagueState:
    rosa: list[int]        # idx dei giocatori della mia squadra
    presi: set[str]        # nomi canonici di tutti i giocatori presi in lega
    spent: float           # crediti spesi dalla mia squadra
    n_acq: int             # acquisti registrati in lega


def build_state(ref: ReferenceData, purchases: list[dict], my_team: str) -> LeagueState:
    """Porting di LeggiAcquisti. `purchases` = lista di dict con chiavi nome/squadra/costo."""
    rosa: list[int] = []
    presi: set[str] = set()
    spent = 0.0
    n_acq = 0
    for p in purchases:
        nome = str(p.get("nome", "")).strip()
        if not nome:
            continue
        n_acq += 1
        idx = ref.by_name.get(nome.lower())
        if idx is None:
            continue
        presi.add(ref.players[idx].nome)
        squadra = str(p.get("squadra", "")).strip()
        if squadra.casefold() == my_team.casefold():
            if len(rosa) < MAX_ROSA:
                rosa.append(idx)
            costo = p.get("costo")
            try:
                spent += float(costo)
            except (TypeError, ValueError):
                pass
    return LeagueState(rosa=rosa, presi=presi, spent=spent, n_acq=n_acq)


# ----------------------------------------------------------------------------
# Classifica moduli, formazione tipo, scarsita', messaggi (porting CalcolaEScrivi)
# ----------------------------------------------------------------------------
@dataclass
class ModuleRank:
    nome: str
    total: float
    covered: int
    uncovered: int
    solution: SlotSolution


@dataclass
class FormationSlot:
    slot_text: str
    player_name: Optional[str]
    squadra_reale: Optional[str]
    role: Optional[str]
    points: float
    covered: bool


@dataclass
class ScarcityRow:
    role: str
    top_name: Optional[str]
    top_rating: float
    liberi: int
    n8: int
    n7: int


@dataclass
class DashboardResult:
    module_ranking: list[ModuleRank]
    best_module: Optional[str]
    formation: list[FormationSlot]
    formation_total: float
    scarcity: list[ScarcityRow]
    messages: list[str]
    credits_left: float
    spent: float
    rosa_size: int
    n_acq: int
    portieri: int


def rank_modules(ref: ReferenceData, rosa: list[int]) -> list[ModuleRank]:
    ranks: list[ModuleRank] = []
    for nome in ref.module_order:
        sol = solve_slots(rosa, ref.modules[nome], ref.players)
        ranks.append(
            ModuleRank(
                nome=nome,
                total=round(sol.total, 2),
                covered=sol.covered,
                uncovered=sol.uncovered,
                solution=sol,
            )
        )
    # ordina per totale desc, poi per slot scoperti asc (come il bubble sort VBA)
    ranks.sort(key=lambda m: (-m.total, m.uncovered))
    return ranks


def _slot_text(module_slots: list[list[int]], s: int) -> str:
    return "/".join(RUOLI[r] for r in module_slots[s])


def _formation_detail(ref: ReferenceData, module_name: str, sol: SlotSolution) -> list[FormationSlot]:
    slots = ref.modules[module_name]
    detail: list[FormationSlot] = []
    for s in range(NSLOT):
        pidx = sol.assignment[s]
        if pidx is None:
            detail.append(FormationSlot(_slot_text(slots, s), None, None, None, 0.0, False))
        else:
            pl = ref.players[pidx]
            role = sol.roles[s]
            detail.append(
                FormationSlot(
                    slot_text=_slot_text(slots, s),
                    player_name=pl.nome,
                    squadra_reale=pl.squadra_reale,
                    role=RUOLI[role] if role is not None else None,
                    points=round(sol.points[s], 2),
                    covered=True,
                )
            )
    return detail


def scarcity(ref: ReferenceData, presi: set[str]) -> list[ScarcityRow]:
    """Porting di ScriviScarsita: per ogni ruolo, i giocatori ancora liberi."""
    rows: list[ScarcityRow] = []
    for j in range(NUM_RUOLI):
        liberi = n8 = n7 = 0
        top_name: Optional[str] = None
        top_rating = -1.0
        for pl in ref.players:
            if pl.fuori_lista:
                continue
            rt = pl.ratings[j]
            if rt < 0:
                continue
            if pl.nome in presi:
                continue
            liberi += 1
            if rt >= 8:
                n8 += 1
            if rt >= 7:
                n7 += 1
            if rt > top_rating:
                top_rating, top_name = rt, f"{pl.nome} ({pl.squadra_reale})"
        rows.append(
            ScarcityRow(
                role=RUOLI[j],
                top_name=top_name,
                top_rating=round(top_rating, 2) if top_name else 0.0,
                liberi=liberi,
                n8=n8,
                n7=n7,
            )
        )
    return rows


def _count_portieri(ref: ReferenceData, rosa: list[int]) -> int:
    return sum(1 for idx in rosa if ref.players[idx].ratings[ROLE_INDEX["Por"]] >= 0)


def _build_messages(
    ref: ReferenceData,
    state: LeagueState,
    ranking: list[ModuleRank],
    formation: list[FormationSlot],
    portieri: int,
) -> list[str]:
    msgs: list[str] = []

    if portieri < MIN_POR:
        msgs.append(f"Portieri: ne mancano {MIN_POR - portieri} (minimo {MIN_POR}).")

    buchi = [f.slot_text for f in formation if not f.covered]
    fully = next((m for m in ranking if m.uncovered == 0), None)
    if buchi:
        msgs.append("Slot scoperti nel modulo di testa: " + ", ".join(buchi) + ".")
        if fully is not None:
            msgs.append(
                f"Miglior modulo interamente coperto: {fully.nome} a {fully.total:.2f}."
            )
        else:
            msgs.append("Nessun modulo e' ancora interamente coperto.")
    else:
        coperti = [f for f in formation if f.covered]
        if coperti:
            weak = min(coperti, key=lambda f: f.points)
            msgs.append(
                f"Slot piu' debole: {weak.slot_text} con {weak.player_name} "
                f"({weak.points:.2f}). Migliorarlo alza subito il totale."
            )

    if len(ranking) > 1:
        second = ranking[1]
        msgs.append(f"Secondo modulo: {second.nome} a {second.total:.2f}.")

    msgs.append(
        f"Crediti residui: {int(CREDITI - state.spent)} su {CREDITI}. "
        f"Giocatori in rosa: {len(state.rosa)}."
    )
    return msgs


def compute_dashboard(ref: ReferenceData, purchases: list[dict], my_team: str) -> DashboardResult:
    """Aggregatore principale: replica l'output della macro `Aggiorna`."""
    state = build_state(ref, purchases, my_team)
    portieri = _count_portieri(ref, state.rosa)
    scar = scarcity(ref, state.presi)

    if not state.rosa:
        return DashboardResult(
            module_ranking=[],
            best_module=None,
            formation=[],
            formation_total=0.0,
            scarcity=scar,
            messages=[
                f"Nessun giocatore in rosa per {my_team}: registra gli acquisti.",
                f"Crediti residui: {int(CREDITI - state.spent)} su {CREDITI}.",
            ],
            credits_left=CREDITI - state.spent,
            spent=state.spent,
            rosa_size=0,
            n_acq=state.n_acq,
            portieri=0,
        )

    ranking = rank_modules(ref, state.rosa)
    best = ranking[0]
    formation = _formation_detail(ref, best.nome, best.solution)
    messages = _build_messages(ref, state, ranking, formation, portieri)

    return DashboardResult(
        module_ranking=ranking,
        best_module=best.nome,
        formation=formation,
        formation_total=round(best.total, 2),
        scarcity=scar,
        messages=messages,
        credits_left=CREDITI - state.spent,
        spent=state.spent,
        rosa_size=len(state.rosa),
        n_acq=state.n_acq,
        portieri=portieri,
    )


# ----------------------------------------------------------------------------
# Campetto: titolari + riserve per slot (porting di DisegnaCampetto)
# ----------------------------------------------------------------------------
@dataclass
class PitchSlot:
    slot_text: str
    starter_name: Optional[str]
    starter_role: Optional[str]
    starter_points: float
    subs: list[tuple[str, float]] = field(default_factory=list)  # (nome, rating)


CP_NSUB = 3


def build_pitch(ref: ReferenceData, rosa: list[int], module_name: str) -> tuple[list[PitchSlot], float, list[str]]:
    """Titolari (assegnamento ottimo) + fino a CP_NSUB riserve per slot, senza ripetizioni.

    Ritorna (slot del campetto, totale rating titolari, giocatori in rosa non impiegabili).
    """
    slots = ref.modules[module_name]
    starters = solve_slots(rosa, slots, ref.players)
    used: set[int] = {a for a in starters.assignment if a is not None}

    sub_map: list[list[tuple[str, float]]] = [[] for _ in range(NSLOT)]
    for _ in range(CP_NSUB):
        panca = [idx for idx in rosa if idx not in used]
        if not panca:
            break
        extra = solve_slots(panca, slots, ref.players)
        found = False
        for s in range(NSLOT):
            pidx = extra.assignment[s]
            if pidx is not None:
                sub_map[s].append((ref.players[pidx].nome, round(extra.points[s], 2)))
                used.add(pidx)
                found = True
        if not found:
            break

    pitch: list[PitchSlot] = []
    for s in range(NSLOT):
        pidx = starters.assignment[s]
        if pidx is None:
            pitch.append(PitchSlot(_slot_text(slots, s), None, None, 0.0, sub_map[s]))
        else:
            pl = ref.players[pidx]
            role = starters.roles[s]
            pitch.append(
                PitchSlot(
                    slot_text=_slot_text(slots, s),
                    starter_name=pl.nome,
                    starter_role=RUOLI[role] if role is not None else None,
                    starter_points=round(starters.points[s], 2),
                    subs=sub_map[s],
                )
            )

    fuori = [ref.players[idx].nome for idx in rosa if idx not in used]
    return pitch, round(starters.total, 2), fuori
